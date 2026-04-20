# -*- coding: utf-8 -*-
"""
Email Scanner - Quét mail Outlook tự động lấy dữ liệu Stock & Bao bì
=====================================================================
- Quét mailbox: phinho@cp.com.vn hoặc mixer2@cp.com.vn
- Tìm email từ: dinhnguyen@cp.com.vn (Tran Dinh Thao Nguyen)
- Tải attachment: FFSTOCK*.xlsm (stock thành phẩm), DAILY STOCK EMPTY BAG*.xlsm (bao bì)
- Parse dữ liệu và gửi lên server: https://teserr.onrender.com/api
"""

import os
import sys
import re
import json
import tempfile
import datetime
import logging
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Fix Windows console encoding
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ========== CONFIG ==========
REMOTE_API = "https://teserr.onrender.com/api"

# Tên thư mục Favorites chứa email cần quét
TARGET_FOLDER_NAME = "Nguyen KTP"

# Người gửi cần tìm (tên hiển thị trong Exchange)
SENDER_NAME = "Tran Dinh Thao Nguyen"
SENDER_EMAIL = "dinhnguyen@cp.com.vn"

# Thư mục lưu attachment tạm
TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")

# Số ngày quét lại
SCAN_DAYS = 7

# ========== LOGGING ==========
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("EmailScanner")


# ========== OUTLOOK CONNECTION ==========

def connect_outlook():
    """Kết nối Outlook COM Object"""
    try:
        import win32com.client
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        log.info("Đã kết nối Outlook thành công")
        return outlook
    except Exception as e:
        log.error(f"Không thể kết nối Outlook: {e}")
        log.error("Hãy đảm bảo Outlook đang chạy và đã đăng nhập")
        return None


def find_folder_recursive(folder, target_name, depth=0, max_depth=5):
    """Tìm folder theo tên, duyệt đệ quy tất cả subfolder"""
    if depth > max_depth:
        return None
    
    try:
        folder_name = folder.Name or ""
        if folder_name.lower().strip() == target_name.lower().strip():
            return folder
        
        for subfolder in folder.Folders:
            found = find_folder_recursive(subfolder, target_name, depth + 1, max_depth)
            if found:
                return found
    except Exception:
        pass
    
    return None


def find_target_folder(outlook):
    """
    Tìm thư mục 'Nguyen KTP' trong tất cả mailbox/store.
    Bỏ qua Online Archive, ưu tiên mailbox chính.
    """
    found_folders = []
    
    for store in outlook.Stores:
        try:
            store_name = store.DisplayName or ""
            
            # Bỏ qua Online Archive (chứa email cũ, không có email mới)
            if "archive" in store_name.lower():
                log.info(f"  Bỏ qua archive: {store_name}")
                continue
            
            root = store.GetRootFolder()
            log.info(f"  Quét store: {store_name}")
            
            result = find_folder_recursive(root, TARGET_FOLDER_NAME, depth=0, max_depth=4)
            if result:
                item_count = 0
                try:
                    item_count = result.Items.Count
                except Exception:
                    pass
                log.info(f"  ✅ Tìm thấy '{TARGET_FOLDER_NAME}' trong [{store_name}] ({item_count} items)")
                found_folders.append((result, store_name, item_count))
        except Exception as e:
            log.debug(f"  Lỗi quét store: {e}")
            continue
    
    if not found_folders:
        log.warning(f"Không tìm thấy thư mục '{TARGET_FOLDER_NAME}' trong bất kỳ mailbox nào")
        return None
    
    # Nếu có nhiều folder, thử quét tất cả (ưu tiên folder có item nhiều nhất)
    # Nhưng cũng quét folder thứ 2 nếu folder đầu không có kết quả
    found_folders.sort(key=lambda x: x[2], reverse=True)
    best = found_folders[0]
    log.info(f"\n📂 Sử dụng: '{TARGET_FOLDER_NAME}' từ [{best[1]}] ({best[2]} items)")
    return found_folders  # Trả về list tất cả folders


def is_sender_match(mail):
    """Kiểm tra email có phải từ sender cần tìm không (hỗ trợ Exchange X500 address)"""
    try:
        # Check SenderName first (reliable cho Exchange)
        sender_name = (mail.SenderName or "").lower()
        if "tran dinh thao" in sender_name or "nguyen ktp" in sender_name:
            return True
        
        # Check SMTP address
        sender_addr = (mail.SenderEmailAddress or "").lower()
        if SENDER_EMAIL.lower() in sender_addr:
            return True
        
        # Check if sender name contains key parts
        if "dinh" in sender_name and "nguyen" in sender_name:
            return True
            
    except Exception:
        pass
    
    return False


def extract_attachments(mail, result):
    """Tải attachment FFSTOCK và DAILY STOCK EMPTY BAG từ email"""
    try:
        att_count = mail.Attachments.Count
        if att_count == 0:
            return False
        
        found_any = False
        for i in range(1, att_count + 1):
            att = mail.Attachments.Item(i)
            filename = att.FileName or ""
            fname_upper = filename.upper()
            
            log.info(f"  📎 {filename} ({att.Size / 1024:.0f} KB)")
            
            # FFSTOCK file
            if "FFSTOCK" in fname_upper and (fname_upper.endswith(".XLSM") or fname_upper.endswith(".XLSX")):
                save_path = os.path.join(TEMP_DIR, filename)
                att.SaveAsFile(save_path)
                result["ffstock_path"] = save_path
                result["email_subject"] = mail.Subject or ""
                result["email_date"] = str(mail.ReceivedTime)
                log.info(f"  ✅ Đã lưu FFSTOCK: {save_path}")
                found_any = True
                
                # Extract date from filename (e.g. "FFSTOCK 19-04-2026.xlsm")
                date_match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
                if date_match:
                    d, m, y = date_match.groups()
                    result["date_str"] = f"{y}-{m}-{d}"
            
            # DAILY STOCK EMPTY BAG file
            if ("DAILY STOCK" in fname_upper or "EMPTY BAG" in fname_upper) and \
               (fname_upper.endswith(".XLSM") or fname_upper.endswith(".XLSX")):
                save_path = os.path.join(TEMP_DIR, filename)
                att.SaveAsFile(save_path)
                result["baobi_path"] = save_path
                log.info(f"  ✅ Đã lưu BaoBi: {save_path}")
                found_any = True
        
        return found_any
    except Exception as e:
        log.debug(f"  Lỗi đọc attachments: {e}")
        return False


def scan_emails(outlook, target_date=None):
    """
    Quét thư mục 'Nguyen KTP' (Favorites) tìm email có attachment FFSTOCK/DAILY STOCK EMPTY BAG.
    Nếu không tìm thấy folder, fallback quét Inbox tất cả mailbox.
    Returns: dict with 'ffstock_path' and 'baobi_path'
    """
    if target_date is None:
        target_date = datetime.date.today()
    
    cutoff = target_date - datetime.timedelta(days=SCAN_DAYS)
    
    os.makedirs(TEMP_DIR, exist_ok=True)
    
    result = {
        "ffstock_path": None,
        "baobi_path": None,
        "date_str": target_date.strftime("%Y-%m-%d"),
        "email_subject": None,
        "email_date": None,
    }
    
    # === STRATEGY 1: Tìm trong thư mục "Nguyen KTP" (Favorites) ===
    log.info(f"\n{'='*50}")
    log.info(f"Tìm thư mục: '{TARGET_FOLDER_NAME}'")
    log.info(f"{'='*50}")
    
    found_folders = find_target_folder(outlook)
    
    if found_folders:
        for folder, store_name, item_count in found_folders:
            log.info(f"\n📂 Quét '{TARGET_FOLDER_NAME}' từ [{store_name}] ({item_count} items)...")
            scan_result = _scan_folder(folder, cutoff, result)
            if scan_result:
                return result
    
    # === STRATEGY 2: Fallback - quét Inbox tìm email từ sender ===
    log.info(f"\n{'='*50}")
    log.info(f"Fallback: Quét Inbox tìm email từ '{SENDER_NAME}'")
    log.info(f"{'='*50}")
    
    for store in outlook.Stores:
        try:
            store_name = store.DisplayName or ""
            root = store.GetRootFolder()
            
            # Find inbox
            inbox = None
            for folder in root.Folders:
                name = folder.Name.lower()
                if name in ("inbox", "hộp thư đến", "hop thu den"):
                    inbox = folder
                    break
            
            if not inbox:
                continue
            
            log.info(f"\nInbox [{store_name}]: {inbox.Items.Count} items")
            scan_result = _scan_folder(inbox, cutoff, result, filter_sender=True)
            if scan_result:
                return result
                
        except Exception:
            continue
    
    return result


def _scan_folder(folder, cutoff, result, filter_sender=False):
    """
    Quét folder tìm email có attachment cần thiết.
    filter_sender=True: chỉ quét email từ sender matching.
    filter_sender=False: quét tất cả email trong folder (vì folder đã lọc sẵn).
    """
    try:
        items = folder.Items
        items.Sort("[ReceivedTime]", True)
        
        # Filter by date
        filter_str = f"[ReceivedTime] >= '{cutoff.strftime('%m/%d/%Y')}'"
        filtered = items.Restrict(filter_str)
        count = filtered.Count
        
        log.info(f"Email từ {cutoff}: {count} email")
        
        if count == 0:
            return False
        
        for mail in filtered:
            try:
                # Skip if filtering by sender and sender doesn't match
                if filter_sender and not is_sender_match(mail):
                    continue
                
                subject = mail.Subject or ""
                sender_name = ""
                try:
                    sender_name = mail.SenderName or ""
                except Exception:
                    pass
                
                # Check if email has attachments
                if not mail.Attachments or mail.Attachments.Count == 0:
                    continue
                
                log.info(f"\n📧 {subject}")
                log.info(f"  Từ: {sender_name}")
                log.info(f"  Ngày: {mail.ReceivedTime}")
                
                # Try to extract attachments
                extract_attachments(mail, result)
                
                # If both found, done!
                if result["ffstock_path"] and result["baobi_path"]:
                    log.info("\n✅ Đã tìm thấy cả 2 file attachment!")
                    return True
                    
            except Exception as e:
                log.debug(f"  Lỗi đọc email: {e}")
                continue
        
        # Partial success
        if result["ffstock_path"] or result["baobi_path"]:
            log.info("\n⚠️ Chỉ tìm thấy 1 trong 2 file")
            return True
            
    except Exception as e:
        log.error(f"Lỗi quét folder: {e}")
    
    return False


# ========== EXCEL PARSERS ==========

def parse_ffstock(filepath):
    """
    Parse file FFSTOCK*.xlsm - Stock thành phẩm
    Sử dụng sheet 'pro' có cấu trúc:
      Col 0: Brand CodeMedicineSize (combined key)
      Col 1: Size (kg)
      Col 2: Brand Code (e.g. '510', '511B', '552F')
      Col 3: Size (kg)
      Col 4: Product Name
      Col 5: Medicine
      Col 11: Quantity (bags) - tồn cuối
      Col 12: Weight (kg) - tồn cuối
    """
    import openpyxl
    
    log.info(f"\n📊 Parsing FFSTOCK: {os.path.basename(filepath)}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"Lỗi mở file: {e}")
        return []
    
    # Use 'pro' sheet
    target_sheet = None
    for name in ['pro', 'Pro', 'PRO']:
        if name in wb.sheetnames:
            target_sheet = name
            break
    
    if not target_sheet:
        log.warning(f"Không tìm thấy sheet 'pro'. Sheets: {wb.sheetnames}")
        # Fallback: try REMIX sheet
        for name in ['REMIX', 'Remix']:
            if name in wb.sheetnames:
                target_sheet = name
                break
    
    if not target_sheet:
        log.error("Không tìm thấy sheet phù hợp trong FFSTOCK")
        wb.close()
        return []
    
    ws = wb[target_sheet]
    log.info(f"Sheet: {target_sheet}")
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log.warning("File rỗng")
        wb.close()
        return []
    
    # Log first rows
    log.info("--- Dữ liệu mẫu ---")
    for i, row in enumerate(rows[:3]):
        log.info(f"  Row {i}: {[str(c)[:25] if c else '' for c in row[:15]]}")
    
    # Parse data: skip header row (row 0), start from row 1
    items = []
    for i in range(1, len(rows)):
        row = rows[i]
        if not row or len(row) < 12:
            continue
        
        # Brand Code at col 2
        code = str(row[2] or "").strip()
        if not code:
            continue
        
        # Skip TOTAL rows
        upper = code.upper()
        if "TOTAL" in upper or "GRAND" in upper:
            continue
        
        # Size at col 3
        pack_size = ""
        try:
            size_val = row[3]
            if size_val:
                pack_size = f"{int(float(size_val))} kg"
        except (ValueError, TypeError):
            pack_size = str(row[3] or "")
        
        # Product Name at col 4
        name = str(row[4] or "").strip()
        
        # Stock Quantity (bags) at col 11
        qty_bags = 0
        try:
            qty_bags = int(float(row[11] or 0))
        except (ValueError, TypeError):
            pass
        
        # Stock Weight (kg) at col 12
        weight_kg = 0
        try:
            weight_kg = float(row[12] or 0)
        except (ValueError, TypeError):
            pass
        
        # Use weight_kg as soLuong (stock in kg)
        stock_kg = weight_kg if weight_kg > 0 else qty_bags * (float(row[3] or 25))
        
        if stock_kg <= 0:
            continue
        
        items.append({
            "codeCam": code,
            "tenCam": name or code,
            "soLuong": stock_kg,
            "packSize": pack_size,
            "balanceBag": float(qty_bags) if qty_bags else None,
            "dayOnHand": None,
            "avgSalePerDay": None,
            "category": "",
        })
    
    wb.close()
    log.info(f"✅ Đã parse {len(items)} sản phẩm từ FFSTOCK")
    
    # Log sample
    for item in items[:5]:
        log.info(f"  {item['codeCam']:10s} | {item['tenCam'][:25]:25s} | {item['soLuong']:>10,.0f} kg | {item.get('balanceBag', 0) or 0:>6,.0f} bags")
    
    return items


def parse_baobi(filepath):
    """
    Parse file DAILY STOCK EMPTY BAG REPORT*.xlsm - Stock bao bì
    Sử dụng sheet 'TOTAL' có cấu trúc:
      Row 2-3: Header (PACKING SIZE, REMAIN, BALANCE...)
      Row 4+: Data:
        Col 0: Brand name (HI-GRO, CP, STAR, NUVO, NASA, FARM)
        Col 1: Packing Size (5, 25, 40, 50)
        Col 8-9: BALANCE (OLD MARKET, NEW MARKET) - Tồn cuối
    """
    import openpyxl
    
    log.info(f"\n📦 Parsing EMPTY BAG: {os.path.basename(filepath)}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"Lỗi mở file: {e}")
        return []
    
    # Use TOTAL sheet first, then individual brand sheets
    target_sheet = None
    for name in ['TOTAL', 'Total']:
        if name in wb.sheetnames:
            target_sheet = name
            break
    
    if not target_sheet:
        log.warning(f"Không tìm thấy sheet 'TOTAL'. Sheets: {wb.sheetnames}")
        wb.close()
        return []
    
    ws = wb[target_sheet]
    log.info(f"Sheet: {target_sheet}")
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log.warning("File rỗng")
        wb.close()
        return []
    
    # Log first rows
    log.info("--- Dữ liệu mẫu ---")
    for i, row in enumerate(rows[:6]):
        log.info(f"  Row {i}: {[str(c)[:25] if c else '' for c in row[:10]]}")
    
    # Data starts at row 4 (index 4)
    # Structure:
    #   Col 0: Brand (HI-GRO BRAND, CP BRAND, etc.) or empty (continuation)
    #   Col 1: Size (5, 25, 40, 50)
    #   Col 8: BALANCE OLD - tồn cuối mẫu cũ
    #   Col 9: BALANCE NEW - tồn cuối mẫu mới
    
    items = []
    current_brand = ""
    
    for i in range(4, len(rows)):
        row = rows[i]
        if not row or len(row) < 10:
            continue
        
        # Get brand name
        brand = str(row[0] or "").strip()
        
        # Skip TOTAL summary rows
        if brand.upper() == "TOTAL":
            break
        
        if brand:
            current_brand = brand.replace("BRAND", "").strip()
        
        if not current_brand:
            continue
        
        # Get size
        size_kg = 25.0
        try:
            if row[1]:
                size_kg = float(row[1])
        except (ValueError, TypeError):
            continue
        
        # Get balance (tồn cuối)
        # Col 8 = OLD MARKET balance, Col 9 = NEW MARKET balance
        balance_old = 0
        balance_new = 0
        try:
            val = row[8]
            if val and str(val) != '#REF!':
                balance_old = int(float(val))
        except (ValueError, TypeError):
            pass
        try:
            val = row[9]
            if val and str(val) != '#REF!':
                balance_new = int(float(val))
        except (ValueError, TypeError):
            pass
        
        total_balance = balance_old + balance_new
        
        if total_balance <= 0:
            continue
        
        loai_bao = f"{current_brand} {int(size_kg)}kg"
        
        items.append({
            "loaiBao": loai_bao,
            "kichCoKg": size_kg,
            "tonKhoHienTai": total_balance,
        })
    
    wb.close()
    log.info(f"✅ Đã parse {len(items)} loại bao bì")
    
    for item in items:
        log.info(f"  {item['loaiBao']:20s} | {item['kichCoKg']:5.0f} kg | {item['tonKhoHienTai']:>10,} cái")
    
    return items


# ========== API SENDER ==========

def send_stock_data(items, date_str, source="email-scanner"):
    """Gửi dữ liệu stock lên server"""
    url = f"{REMOTE_API}/import/stock-scan"
    payload = {
        "date": date_str,
        "source": source,
        "items": items,
    }
    
    log.info(f"\n🚀 Gửi {len(items)} sản phẩm stock lên {url}")
    
    try:
        resp = requests.post(url, json=payload, timeout=60, verify=False)
        log.info(f"  HTTP {resp.status_code}")
        try:
            data = resp.json()
            log.info(f"  Response: {data.get('message', data)}")
            if data.get("newProducts"):
                log.info(f"  Sản phẩm mới: {data['newProducts']}")
            return data
        except Exception:
            log.info(f"  Response (text): {resp.text[:200]}")
            return {"status": resp.status_code, "text": resp.text[:200]}
    except Exception as e:
        log.error(f"  Lỗi gửi stock: {e}")
        return None


def send_baobi_data(items, date_str, source="email-scanner"):
    """Gửi dữ liệu bao bì lên server"""
    url = f"{REMOTE_API}/import/baobi-scan"
    payload = {
        "date": date_str,
        "source": source,
        "items": items,
    }
    
    log.info(f"\n🚀 Gửi {len(items)} loại bao bì lên {url}")
    
    try:
        resp = requests.post(url, json=payload, timeout=60, verify=False)
        log.info(f"  HTTP {resp.status_code}")
        try:
            data = resp.json()
            log.info(f"  Response: {data.get('message', data)}")
            return data
        except Exception:
            log.info(f"  Response (text): {resp.text[:200]}")
            return {"status": resp.status_code, "text": resp.text[:200]}
    except Exception as e:
        log.error(f"  Lỗi gửi bao bì: {e}")
        return None


# ========== MAIN ==========

def main():
    print()
    print("=" * 60)
    print("  EMAIL SCANNER - Quét Stock & Bao bì từ Outlook")
    print("=" * 60)
    print(f"  Server:  {REMOTE_API}")
    print(f"  Folder:  {TARGET_FOLDER_NAME}")
    print(f"  Sender:  {SENDER_NAME} ({SENDER_EMAIL})")
    print(f"  Quét:    {SCAN_DAYS} ngày gần nhất")
    print("=" * 60)
    print()
    
    # 1. Connect Outlook
    outlook = connect_outlook()
    if not outlook:
        return
    
    # 2. Scan emails
    log.info("Bắt đầu quét email...")
    result = scan_emails(outlook)
    
    date_str = result["date_str"]
    source = f"Email: {result.get('email_subject', 'N/A')}"
    
    # 3. Parse & Send FFSTOCK
    if result["ffstock_path"]:
        items = parse_ffstock(result["ffstock_path"])
        if items:
            send_stock_data(items, date_str, source)
        else:
            log.warning("Không parse được dữ liệu từ FFSTOCK")
    else:
        log.warning("❌ Không tìm thấy file FFSTOCK trong email")
    
    # 4. Parse & Send BaoBi
    if result["baobi_path"]:
        items = parse_baobi(result["baobi_path"])
        if items:
            send_baobi_data(items, date_str, source)
        else:
            log.warning("Không parse được dữ liệu từ DAILY STOCK EMPTY BAG")
    else:
        log.warning("❌ Không tìm thấy file DAILY STOCK EMPTY BAG trong email")
    
    # Summary
    print()
    print("=" * 60)
    print("  KẾT QUẢ")
    print("=" * 60)
    print(f"  Ngày:     {date_str}")
    print(f"  FFSTOCK:  {'✅ Đã gửi' if result['ffstock_path'] else '❌ Không tìm thấy'}")
    print(f"  Bao bì:   {'✅ Đã gửi' if result['baobi_path'] else '❌ Không tìm thấy'}")
    print("=" * 60)
    print()


if __name__ == "__main__":
    main()
